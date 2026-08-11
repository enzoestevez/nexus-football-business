from openpyxl import load_workbook

path = "/home/ubuntu/nexus-futbol-business/planilla_operativa_nexus.xlsx"
wb = load_workbook(path)
print("OK")
print("Sheets:", ", ".join(wb.sheetnames))
for name in wb.sheetnames:
    ws = wb[name]
    print(name, ws.max_row, ws.max_column)
