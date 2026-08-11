from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUTPUT = "/home/ubuntu/nexus-futbol-business/planilla_operativa_nexus.xlsx"

wb = Workbook()

# Remove default sheet
wb.remove(wb.active)

# Theme
navy_fill = PatternFill("solid", fgColor="0B2E59")
blue_fill = PatternFill("solid", fgColor="1E63B4")
light_fill = PatternFill("solid", fgColor="EAF2FF")
section_fill = PatternFill("solid", fgColor="D9E8FF")
white_font = Font(color="FFFFFF", bold=True)
header_font = Font(color="FFFFFF", bold=True)
subheader_font = Font(color="0B2E59", bold=True)
thin = Side(style="thin", color="C7D2E3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

lists = {
    "prioridad": ["Alta", "Media", "Baja"],
    "responsable": [
        "Alberto",
        "Esposa / Legal-Administración",
        "Reclutador",
        "Preparador físico",
        "Psicólogo",
        "Diseño / Flyers",
        "Importadores / Cascos",
        "Equipo Nexus",
    ],
    "estado_jugador": [
        "Nuevo",
        "Contactado",
        "Primer llamado",
        "Reunión agendada",
        "Reunión realizada",
        "Documentación solicitada",
        "Diagnóstico ofrecido",
        "Propuesta enviada",
        "Negociación",
        "Cierre ganado",
        "Cierre perdido",
        "En pausa",
    ],
    "categoria_pipeline": ["A", "B", "C"],
    "estado_club": [
        "Nuevo",
        "Contacto inicial",
        "Reunión agendada",
        "Reunión realizada",
        "Diagnóstico propuesto",
        "Propuesta enviada",
        "Negociación",
        "Retainer activo",
        "Proyecto activo",
        "Cierre perdido",
        "En pausa",
    ],
    "tipo_contacto_institucional": ["Club", "Dirigente", "Institución", "Academia", "Liga", "Otro"],
    "estado_casco": [
        "Idea",
        "Cotización",
        "Muestra",
        "Diseño",
        "Costo validado",
        "Preventa",
        "Financiación en análisis",
        "Producción",
        "Listo para venta",
        "Pausado",
    ],
    "financiacion": ["Preventa", "Inversor", "Capital propio", "Mixto", "Sin definir"],
    "estado_inversor": [
        "Nuevo",
        "Contacto inicial",
        "Interés detectado",
        "Reunión agendada",
        "Reunión realizada",
        "Información enviada",
        "Negociación",
        "Compromiso logrado",
        "Descartado",
        "En pausa",
    ],
    "estado_proveedor": [
        "Nuevo",
        "Cotización recibida",
        "Validación técnica",
        "Negociación",
        "Aprobado",
        "Muestra solicitada",
        "Activo",
        "Descartado",
    ],
    "tipo_proveedor": ["Fabricante", "Importador", "Logística", "Packaging", "Diseño", "Otro"],
    "estado_tarea": ["Pendiente", "En curso", "Esperando tercero", "Completada", "Cancelada"],
    "area": ["Futbolistas", "Clubes", "Turismo", "Cascos", "Legal-Administración", "Marketing", "Dirección"],
}

list_sheet = wb.create_sheet("Listas")
list_sheet.sheet_state = "hidden"
for col_idx, (key, values) in enumerate(lists.items(), start=1):
    list_sheet.cell(row=1, column=col_idx, value=key)
    for row_idx, value in enumerate(values, start=2):
        list_sheet.cell(row=row_idx, column=col_idx, value=value)

list_ranges = {}
for col_idx, (key, values) in enumerate(lists.items(), start=1):
    letter = get_column_letter(col_idx)
    list_ranges[key] = f"'Listas'!${letter}$2:${letter}${len(values)+1}"


def style_headers(ws, row=1):
    for cell in ws[row]:
        cell.fill = navy_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center
    ws.row_dimensions[row].height = 26


def set_col_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def apply_table_style(ws, start_row=2, end_row=300):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"
    for row in ws.iter_rows(min_row=2, max_row=end_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = left_wrap


def add_dropdown(ws, col, formula, start_row=2, end_row=300):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}{start_row}:{col}{end_row}")


# Instrucciones
ws = wb.create_sheet("Instrucciones")
ws["A1"] = "Planilla operativa de Nexus Football & Business"
ws["A1"].fill = navy_fill
ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
ws["A1"].alignment = left_wrap
ws.merge_cells("A1:F1")

instructions = [
    ["Objetivo", "Usar esta planilla como pipeline obligatorio para que ningún jugador, club, proveedor o negocio quede solo en WhatsApp o en la memoria."],
    ["Cómo usarla", "Cada oportunidad debe tener responsable, estado, prioridad, próximo paso y fecha. Si no tiene próximo paso, no está siendo gestionada."],
    ["Hojas", "Jugadores | Clubes | Cascos | Inversores | Proveedores | Tareas semanales | Resumen"],
    ["Regla central", "Actualizar la planilla al menos una vez por semana, idealmente antes de la reunión de dirección de Nexus."],
    ["Sugerencia", "Clasificá rápido: A = alta chance / B = seguimiento / C = maduración. Priorizá caja y cierre, no solo volumen."],
]
for r, row in enumerate(instructions, start=3):
    ws.cell(row=r, column=1, value=row[0]).fill = section_fill
    ws.cell(row=r, column=1).font = subheader_font
    ws.cell(row=r, column=2, value=row[1])
    for c in range(1, 3):
        ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=c).alignment = left_wrap
set_col_widths(ws, [22, 110, 16, 16, 16, 16])

# Resumen
ws = wb.create_sheet("Resumen")
summary_headers = ["Indicador", "Valor"]
ws.append(summary_headers)
style_headers(ws)
summary_rows = [
    ["Jugadores activos", "=COUNTA(Jugadores!A2:A300)"],
    ["Jugadores A", '=COUNTIF(Jugadores!J:J,"A")'],
    ["Clubes / dirigentes activos", "=COUNTA(Clubes!A2:A300)"],
    ["Propuestas a clubes enviadas", '=COUNTIF(Clubes!K:K,"Propuesta enviada")'],
    ["Cascos en preventa / producción", '=COUNTIF(Cascos!K:K,"Preventa")+COUNTIF(Cascos!K:K,"Producción")'],
    ["Inversores en negociación", '=COUNTIF(Inversores!G:G,"Negociación")'],
    ["Monto potencial jugadores (USD)", "=SUM(Jugadores!P:P)"],
    ["Monto potencial clubes (USD)", "=SUM(Clubes!L:L)"],
]
for row in summary_rows:
    ws.append(row)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.border = border
        cell.alignment = left_wrap
set_col_widths(ws, [40, 24])

# Jugadores
ws = wb.create_sheet("Jugadores")
player_headers = [
    "ID",
    "Fecha ingreso",
    "Nombre y apellido",
    "Edad",
    "Posición",
    "País / Ciudad",
    "Club actual",
    "Situación contractual",
    "Necesidad principal",
    "Categoría pipeline",
    "Estado comercial",
    "Prioridad",
    "Responsable Nexus",
    "Fuente del contacto",
    "Probabilidad cierre (%)",
    "Potencial ingreso USD",
    "Próximo paso",
    "Fecha próximo paso",
    "Último contacto",
    "Observaciones",
]
ws.append(player_headers)
style_headers(ws)
set_col_widths(ws, [10, 14, 28, 10, 16, 18, 22, 22, 26, 14, 22, 12, 20, 18, 18, 18, 26, 16, 16, 30])
apply_table_style(ws)
add_dropdown(ws, "J", f"={list_ranges['categoria_pipeline']}")
add_dropdown(ws, "K", f"={list_ranges['estado_jugador']}")
add_dropdown(ws, "L", f"={list_ranges['prioridad']}")
add_dropdown(ws, "M", f"={list_ranges['responsable']}")

# Clubes
ws = wb.create_sheet("Clubes")
club_headers = [
    "ID",
    "Fecha ingreso",
    "Tipo contacto",
    "Institución",
    "Nombre del contacto",
    "Cargo",
    "Teléfono",
    "Email",
    "Necesidad detectada",
    "Servicio / puerta de entrada",
    "Estado comercial",
    "Ticket estimado USD",
    "Prioridad",
    "Responsable Nexus",
    "Próximo paso",
    "Fecha próximo paso",
    "Último contacto",
    "Observaciones",
]
ws.append(club_headers)
style_headers(ws)
set_col_widths(ws, [10, 14, 16, 24, 24, 18, 18, 26, 28, 26, 22, 16, 12, 20, 26, 16, 16, 32])
apply_table_style(ws)
add_dropdown(ws, "C", f"={list_ranges['tipo_contacto_institucional']}")
add_dropdown(ws, "K", f"={list_ranges['estado_club']}")
add_dropdown(ws, "M", f"={list_ranges['prioridad']}")
add_dropdown(ws, "N", f"={list_ranges['responsable']}")

# Cascos
ws = wb.create_sheet("Cascos")
helmet_headers = [
    "ID",
    "Fecha registro",
    "Modelo / Club",
    "Proveedor principal",
    "MOQ",
    "Costo unitario USD",
    "Costos extras USD",
    "Costo total unitario USD",
    "Precio objetivo USD",
    "Margen unitario USD",
    "Estado",
    "Financiación",
    "Responsable",
    "Próximo paso",
    "Fecha próximo paso",
    "Observaciones",
]
ws.append(helmet_headers)
style_headers(ws)
set_col_widths(ws, [10, 14, 24, 24, 10, 16, 16, 18, 16, 16, 22, 18, 20, 26, 16, 32])
apply_table_style(ws)
add_dropdown(ws, "K", f"={list_ranges['estado_casco']}")
add_dropdown(ws, "L", f"={list_ranges['financiacion']}")
add_dropdown(ws, "M", f"={list_ranges['responsable']}")

# Inversores
ws = wb.create_sheet("Inversores")
investor_headers = [
    "ID",
    "Fecha ingreso",
    "Nombre",
    "Tipo inversor",
    "Interés potencial",
    "Monto estimado USD",
    "Estado",
    "Prioridad",
    "Responsable",
    "Próximo paso",
    "Fecha próximo paso",
    "Observaciones",
]
ws.append(investor_headers)
style_headers(ws)
set_col_widths(ws, [10, 14, 26, 18, 28, 18, 22, 12, 20, 24, 16, 32])
apply_table_style(ws)
add_dropdown(ws, "G", f"={list_ranges['estado_inversor']}")
add_dropdown(ws, "H", f"={list_ranges['prioridad']}")
add_dropdown(ws, "I", f"={list_ranges['responsable']}")

# Proveedores
ws = wb.create_sheet("Proveedores")
supplier_headers = [
    "ID",
    "Fecha ingreso",
    "Proveedor",
    "Tipo",
    "Producto / servicio",
    "Contacto",
    "Teléfono / Email",
    "Costo estimado",
    "Estado",
    "Prioridad",
    "Responsable",
    "Próximo paso",
    "Fecha próximo paso",
    "Observaciones",
]
ws.append(supplier_headers)
style_headers(ws)
set_col_widths(ws, [10, 14, 24, 18, 24, 22, 24, 16, 20, 12, 20, 24, 16, 30])
apply_table_style(ws)
add_dropdown(ws, "D", f"={list_ranges['tipo_proveedor']}")
add_dropdown(ws, "I", f"={list_ranges['estado_proveedor']}")
add_dropdown(ws, "J", f"={list_ranges['prioridad']}")
add_dropdown(ws, "K", f"={list_ranges['responsable']}")

# Tareas semanales
ws = wb.create_sheet("Tareas semanales")
task_headers = [
    "ID",
    "Semana",
    "Área",
    "Tarea",
    "Responsable",
    "Estado",
    "Fecha compromiso",
    "Resultado esperado",
    "Observaciones",
]
ws.append(task_headers)
style_headers(ws)
set_col_widths(ws, [10, 14, 18, 34, 20, 18, 16, 26, 30])
apply_table_style(ws)
add_dropdown(ws, "C", f"={list_ranges['area']}")
add_dropdown(ws, "E", f"={list_ranges['responsable']}")
add_dropdown(ws, "F", f"={list_ranges['estado_tarea']}")

for sheet in wb.worksheets:
    if sheet.title != "Listas":
        sheet.sheet_view.showGridLines = True

wb.save(OUTPUT)
print(OUTPUT)
